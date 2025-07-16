import os
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, date

def apply_date_format_to_column(ws, col_index, format_string):
    col_letter = get_column_letter(col_index)
    for row in range(2, ws.max_row + 1):  # Skip header
        cell = ws[f"{col_letter}{row}"]
        if isinstance(cell.value, (datetime, date)):
            cell.number_format = format_string

def update_xlsm_with_bhavcopy(xlsm_file: str):
    # === CONFIGURATION ===
    download_dir = os.getcwd()
    csv_prefix = "BhavCopyDateWise_"

    # === FIND LATEST CSV FILE ===
    csv_files = [f for f in os.listdir(download_dir) if f.startswith(csv_prefix) and f.endswith(".csv")]
    if not csv_files:
        raise FileNotFoundError("❌ No BhavCopyDateWise_*.csv file found in the directory.")

    latest_csv = max(csv_files, key=lambda f: os.path.getmtime(os.path.join(download_dir, f)))
    csv_path = os.path.join(download_dir, latest_csv)
    print(f"📄 Using latest CSV: {latest_csv}")

    # === LOAD CSV FILE ===
    df_csv = pd.read_csv(csv_path)
    df_csv.columns = df_csv.columns.str.strip()

    # Normalize date columns
    df_csv["Date"] = pd.to_datetime(df_csv["Date"], format="%d %b %Y")
    df_csv["Expiry Date"] = pd.to_datetime(df_csv["Expiry Date"], format="%d%b%Y")

    # Strip strings
    df_csv["Instrument Name"] = df_csv["Instrument Name"].astype(str).str.strip()
    df_csv["Symbol"] = df_csv["Symbol"].astype(str).str.strip()

    # === LOAD .XLSM FILE (PRESERVE MACROS) ===
    wb = openpyxl.load_workbook(xlsm_file, keep_vba=True)

    # === PROCESS FUTURES SHEET ===
    futures_ws = wb["Futures"]
    f_instr = str(futures_ws["B2"].value).strip()
    f_sym = str(futures_ws["C2"].value).strip()
    f_exp_raw = futures_ws["D2"].value
    f_exp = f_exp_raw.date() if isinstance(f_exp_raw, datetime) else f_exp_raw

    df_futures = df_csv[
        (df_csv["Instrument Name"] == f_instr) &
        (df_csv["Symbol"] == f_sym) &
        (df_csv["Expiry Date"].dt.date == f_exp)
    ]

    if not df_futures.empty:
        print(f"📊 Appending {len(df_futures)} futures rows...")
        for row in df_futures.itertuples(index=False):
            futures_ws.append(list(row))
    else:
        print("⚠️ No matching futures rows found.")

    # Format Date (A) and Expiry Date (D) in Futures sheet
    apply_date_format_to_column(futures_ws, 1, "DD/MM/YYYY")
    apply_date_format_to_column(futures_ws, 4, "DD/MM/YYYY")

    # === PROCESS OPTIONS SHEET ===
    options_ws = wb["Options"]
    o_instr = str(options_ws["B2"].value).strip()
    o_sym = str(options_ws["C2"].value).strip()
    o_exp_raw = options_ws["D2"].value
    o_exp = o_exp_raw.date() if isinstance(o_exp_raw, datetime) else o_exp_raw

    df_options = df_csv[
        (df_csv["Instrument Name"] == o_instr) &
        (df_csv["Symbol"] == o_sym) &
        (df_csv["Expiry Date"].dt.date == o_exp)
    ]

    if not df_options.empty:
        print(f"📊 Appending {len(df_options)} options rows...")
        for row in df_options.itertuples(index=False):
            options_ws.append(list(row))
    else:
        print("⚠️ No matching options rows found.")

    # Format Date (A) and Expiry Date (D) in Options sheet
    apply_date_format_to_column(options_ws, 1, "DD/MM/YYYY")
    apply_date_format_to_column(options_ws, 4, "DD/MM/YYYY")

    # === SAVE .XLSM FILE ===
    wb.save(xlsm_file)
    print(f"✅ Updated workbook: {xlsm_file}")


# Example usage
if __name__ == "__main__":
    update_xlsm_with_bhavcopy("GOLD option chain.xlsm")
