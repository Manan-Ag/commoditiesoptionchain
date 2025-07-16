import os
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, date

def add_date_dropdown_to_option_chain(xlsm_file: str):
    # Load workbook
    wb = openpyxl.load_workbook(xlsm_file, keep_vba=True)

    # Reference sheets
    futures_ws = wb["Futures"]
    option_chain_ws = wb["Option chain"]

    # === Step 1: Extract unique dates from column A of Futures sheet ===
    unique_dates = set()
    for cell in futures_ws["A"][1:]:  # skip header
        val = cell.value
        if val:
            if isinstance(val, (datetime, date)):
                unique_dates.add(val.date())
            else:
                try:
                    parsed = datetime.strptime(str(val), "%d/%m/%Y").date()
                    unique_dates.add(parsed)
                except Exception:
                    continue

    sorted_dates = sorted(unique_dates)

    # === Step 2: Create or refresh Validation sheet ===
    if "Validation" in wb.sheetnames:
        val_ws = wb["Validation"]
        # Clear A1:A100 only
        for i in range(1, 101):
            val_ws.cell(row=i, column=1).value = None
    else:
        val_ws = wb.create_sheet("Validation")
        val_ws.sheet_state = "hidden"  # hide it

    # Write dates into Validation sheet
    for i, d in enumerate(sorted_dates[:100], start=1):  # max 100 dates
        val_ws.cell(row=i, column=1, value=d)
        val_ws.cell(row=i, column=1).number_format = "DD/MM/YYYY"

    # === Step 3: Add data validation to Option chain!B1 ===
    dv = DataValidation(
        type="list",
        formula1="=Validation!$A$1:$A$100",
        allow_blank=True,
        showDropDown=False
    )
    dv.prompt = "Choose a date from the dropdown"
    dv.error = "Invalid date selected"

    option_chain_ws.add_data_validation(dv)
    dv.add(option_chain_ws["B1"])

    # === Save the file ===
    wb.save(xlsm_file)
    print(f"✅ Dropdown added to Option chain!B1 from Validation!A1:A100 → {xlsm_file}")


# Example usage
if __name__ == "__main__":
    add_date_dropdown_to_option_chain("GOLD option chain.xlsm")
